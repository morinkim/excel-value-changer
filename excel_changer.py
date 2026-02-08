# -*- coding: utf-8 -*-
"""
Excel Value Changer
- Batch update specific cell values across all Excel files (.xlsx) in the data folder
"""

import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
from openpyxl import load_workbook
import threading


class ExcelChangerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Value Changer")
        self.root.geometry("500x450")
        self.root.resizable(False, False)

        # Set data folder path based on executable location
        if getattr(sys, 'frozen', False):
            # Running as exe
            self.base_path = os.path.dirname(sys.executable)
        else:
            # Running as Python script
            self.base_path = os.path.dirname(os.path.abspath(__file__))

        self.data_folder = os.path.join(self.base_path, "data")

        self.create_widgets()
        self.check_data_folder()

    def create_widgets(self):
        # Info section
        info_frame = ttk.LabelFrame(self.root, text="Instructions", padding=10)
        info_frame.pack(fill="x", padx=10, pady=5)

        info_text = f"Data folder: {self.data_folder}\n"
        info_text += "Place Excel files (.xlsx) in the folder above, then enter the cells and value below."
        ttk.Label(info_frame, text=info_text, wraplength=460).pack()

        # Cell input section
        cell_frame = ttk.LabelFrame(self.root, text="Cell Settings", padding=10)
        cell_frame.pack(fill="x", padx=10, pady=5)

        # Cell address input
        ttk.Label(cell_frame, text="Cell addresses (comma-separated):").pack(anchor="w")
        ttk.Label(cell_frame, text="e.g. A1, B2, C3", foreground="gray").pack(anchor="w")
        self.cell_entry = ttk.Entry(cell_frame, width=50)
        self.cell_entry.pack(fill="x", pady=(0, 10))
        self.cell_entry.insert(0, "A1")

        # New value input
        ttk.Label(cell_frame, text="New value:").pack(anchor="w")
        self.value_entry = ttk.Entry(cell_frame, width=50)
        self.value_entry.pack(fill="x")

        # Buttons
        btn_frame = ttk.Frame(self.root, padding=10)
        btn_frame.pack(fill="x")

        self.run_btn = ttk.Button(btn_frame, text="Run Batch Change", command=self.run_change)
        self.run_btn.pack(side="left", padx=5)

        self.open_folder_btn = ttk.Button(btn_frame, text="Open Data Folder", command=self.open_data_folder)
        self.open_folder_btn.pack(side="left", padx=5)

        # Log section
        log_frame = ttk.LabelFrame(self.root, text="Results", padding=10)
        log_frame.pack(fill="both", expand=True, padx=10, pady=5)

        self.log_text = scrolledtext.ScrolledText(log_frame, height=10, state="disabled")
        self.log_text.pack(fill="both", expand=True)

    def check_data_folder(self):
        """Create data folder if it doesn't exist"""
        if not os.path.exists(self.data_folder):
            os.makedirs(self.data_folder)
            self.log(f"Created data folder: {self.data_folder}")
        else:
            files = [f for f in os.listdir(self.data_folder) if f.endswith(('.xlsx', '.xls'))]
            self.log(f"Found {len(files)} Excel file(s) in data folder")

    def log(self, message):
        """Add log message"""
        self.log_text.config(state="normal")
        self.log_text.insert("end", message + "\n")
        self.log_text.see("end")
        self.log_text.config(state="disabled")

    def clear_log(self):
        """Clear log"""
        self.log_text.config(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.config(state="disabled")

    def open_data_folder(self):
        """Open data folder in explorer"""
        if not os.path.exists(self.data_folder):
            os.makedirs(self.data_folder)
        os.startfile(self.data_folder)

    def parse_cells(self, cell_string):
        """Parse cell address string into list"""
        cells = []
        for cell in cell_string.split(","):
            cell = cell.strip().upper()
            if cell:
                cells.append(cell)
        return cells

    def run_change(self):
        """Run batch change"""
        cells_str = self.cell_entry.get().strip()
        new_value = self.value_entry.get()

        if not cells_str:
            messagebox.showwarning("Input Error", "Please enter cell address(es).")
            return

        cells = self.parse_cells(cells_str)
        if not cells:
            messagebox.showwarning("Input Error", "Please enter valid cell address(es).")
            return

        # Check data folder
        if not os.path.exists(self.data_folder):
            messagebox.showwarning("Folder Not Found", "Data folder not found. Creating folder.")
            os.makedirs(self.data_folder)
            return

        # Get Excel file list
        excel_files = [f for f in os.listdir(self.data_folder) if f.endswith(('.xlsx',))]

        if not excel_files:
            messagebox.showwarning("No Files", "No Excel files (.xlsx) found in data folder.")
            return

        # Confirmation dialog
        confirm_msg = f"The following action will be performed:\n\n"
        confirm_msg += f"Target files: {len(excel_files)}\n"
        confirm_msg += f"Cells to change: {', '.join(cells)}\n"
        confirm_msg += f"New value: '{new_value}'\n\n"
        confirm_msg += "Do you want to continue?"

        if not messagebox.askyesno("Confirm", confirm_msg):
            return

        # Disable button
        self.run_btn.config(state="disabled")
        self.clear_log()

        # Process in separate thread
        thread = threading.Thread(target=self.process_files, args=(excel_files, cells, new_value))
        thread.start()

    def process_files(self, excel_files, cells, new_value):
        """Process Excel files"""
        success_count = 0
        fail_count = 0

        for filename in excel_files:
            filepath = os.path.join(self.data_folder, filename)
            try:
                self.log(f"Processing: {filename}")

                # Open Excel file
                wb = load_workbook(filepath)
                ws = wb.active  # First sheet

                # Change cell values
                for cell in cells:
                    ws[cell] = new_value

                # Save
                wb.save(filepath)
                wb.close()

                self.log(f"  Done: {', '.join(cells)} -> '{new_value}'")
                success_count += 1

            except Exception as e:
                self.log(f"  Error: {str(e)}")
                fail_count += 1

        # Summary
        self.log("-" * 40)
        self.log(f"Complete: {success_count} succeeded, {fail_count} failed")

        # Re-enable button
        self.root.after(0, lambda: self.run_btn.config(state="normal"))

        # Completion message
        self.root.after(0, lambda: messagebox.showinfo(
            "Complete",
            f"Processing complete.\nSucceeded: {success_count}\nFailed: {fail_count}"
        ))


def main():
    root = tk.Tk()
    app = ExcelChangerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
